import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

wb: Workbook = openpyxl.load_workbook('a.xlsx')
sheet_names = wb.sheetnames
st:Worksheet=wb.active
print(type(st))
st.cell(1,1,"")
wb.save('a1.xlsx')
