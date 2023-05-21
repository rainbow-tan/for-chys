import csv
import os

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def traverse_folder(path: str, n: int = 9999):
    path = os.path.abspath(path)
    assert os.path.isdir(path)
    assert n > 0
    all_files = []
    all_dirs = []
    for root, dirs, files in os.walk(path):
        for one_file in files:
            all_files.append(os.path.join(root, one_file))  # 所有文件
        for one_dir in dirs:
            all_dirs.append(os.path.join(root, one_dir))  # 所有文件夹

    path_split = path.split(os.sep)
    len_path_split = len(path_split)

    need_dir = []
    for d_name in all_dirs:
        dir_split = d_name.split(os.sep)
        dir_split_ = dir_split[len_path_split:]
        if len(dir_split_) <= n:
            need_dir.append(d_name)

    need_files = []
    for f_name in all_files:
        f_name_split = f_name.split(os.sep)
        f_name_split_ = f_name_split[len_path_split:]
        if len(f_name_split_) <= n:
            need_files.append(f_name)

    return need_dir, need_files


def find_csvs() -> dict:
    _, files = traverse_folder(".")
    csvs = list(filter(lambda x: str(x).lower().endswith('.csv'), files))
    print(f"csvs:{csvs}")
    csv_info = dict()
    for csv_name in csvs:
        basename = os.path.basename(csv_name)
        for_me = basename.split("-")[0]
        csv_info[for_me] = csv_name
    print(f"csv info:{csv_info}")
    return csv_info


def read_csv_content(filename):
    src_data = []
    with open(filename) as f:
        reader = csv.reader(f)
        for row in reader:
            src_data.append(row)
    print(f"src data:{src_data}")
    start = 'A'  # 改这两个地方
    end = 'Z'  # 改这两个地方
    need_data = []
    for row in src_data:
        need = []
        for i in range(ord(start), ord(end) + 1):
            if i - 65 < len(row):
                need.append(row[i - 65])
            else:
                need.append("")
        need_data.append(need)
    print(f"need    :{need_data}")

    return need_data


def read_all_csv(csv_info: dict) -> dict:
    csv_data = dict()
    for k, v in csv_info.items():
        need_data = read_csv_content(v)
        csv_data[k] = need_data
    return csv_data


XLSX_NAME = "XLSX_File_Before_RunCode.xlsx"


def update_xlsx(csv_data: dict):
    wb: Workbook = openpyxl.load_workbook(XLSX_NAME)
    sheet_names = wb.sheetnames
    print(f"sheet names:{sheet_names}")
    sheet_info = dict()
    for sheet_name in sheet_names:
        for_me = sheet_name.split('-')[0]
        sheet_info[for_me] = sheet_name
    print(f"sheet info:{sheet_info}")

    for k, v in sheet_info.items():
        i = 1
        j = 1
        sheet: Worksheet = wb[v]
        insert_data = csv_data[k]
        for row in insert_data:
            for data in row:
                if not data:
                    continue
                sheet.cell(i, j, data)
                j += 1
            i += 1
            j = 1
    wb.save(XLSX_NAME)


def main():
    csv_info = find_csvs()
    csv_data = read_all_csv(csv_info)
    update_xlsx(csv_data)


if __name__ == '__main__':
    main()
