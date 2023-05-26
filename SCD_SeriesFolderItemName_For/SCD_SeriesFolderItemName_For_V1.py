import os.path
from typing import List

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def read_xlsx(filename: str, sheet_name: str) -> List[list]:
    filename = os.path.abspath(filename)
    assert os.path.isfile(filename)
    wb: Workbook = openpyxl.load_workbook(filename)
    assert sheet_name in wb.sheetnames, "sheet name error"
    ws: Worksheet = wb[sheet_name]
    all_data = []
    for row in ws.rows:
        data_row = []
        for data in row:
            cell: Cell = data
            data_row.append(cell.value)
        all_data.append(data_row)
    print("all data".center(50, '*'))
    for one in all_data:
        print(one)
    return all_data


def deal_with_all_data(all_data: List[list]):
    prefix = None
    for row in all_data:
        a_value = row[0]
        if 'Folder' in a_value:
            e_value = row[4]
            prefix = e_value
        else:
            if prefix:
                e = row[4]
                e_after = f'{prefix}-{e}'
                row[4] = e_after

    print("after data".center(50, '*'))
    for one in all_data:
        print(one)
    return all_data


def save_file(filename: str, sheet_name: str, all_data: List[list]):
    filename = os.path.abspath(filename)
    assert os.path.isfile(filename)
    wb: Workbook = openpyxl.load_workbook(filename)
    assert sheet_name in wb.sheetnames, "sheet name error"
    ws: Worksheet = wb[sheet_name]

    for index, row in enumerate(all_data):
        for index2, cell in enumerate(row):
            ws.cell(index + 1, index2 + 1, cell)
    wb.save(filename)


def main():
    filename = "../../learn-python/读取xls/SCD_SeriesFolderItemName_ForV1_Before.xls"
    sheet_name = "Sheet1"
    all_data = read_xlsx(filename, sheet_name)
    all_data = deal_with_all_data(all_data)
    save_file(filename, sheet_name, all_data)


if __name__ == '__main__':
    main()
